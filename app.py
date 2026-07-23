import io
import re
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


@dataclass
class Record:
    name: str
    employee_id: str
    title: str
    revision: str


NAME_PATTERNS = [
    re.compile(r"(?:trainee\s*name|name)\s*[:\-]\s*([A-Za-z][A-Za-z\s,.'-]{1,80})", re.IGNORECASE),
    re.compile(r"employee\s*name\s*[:\-]\s*([A-Za-z][A-Za-z\s,.'-]{1,80})", re.IGNORECASE),
]

EMPLOYEE_ID_PATTERNS = [
    re.compile(r"(?:trainee\s*employee\s*id|employee\s*id|emp\s*id)\s*[:\-]\s*([A-Za-z0-9\-/]{2,30})", re.IGNORECASE),
    re.compile(r"\bID\s*[:\-]\s*([A-Za-z0-9\-/]{2,30})\b", re.IGNORECASE),
]

# Captures pairs such as:
# DOC-123 Rev 01, Safety Basics
# SOP-21 / 02 Lockout Tagout
DOC_REV_PATTERNS = [
    re.compile(
        r"(?P<doc>[A-Za-z]{1,10}[A-Za-z0-9\-_/]{1,30})\s*(?:,|\-|/)?\s*(?:rev(?:ision)?\s*)?(?P<rev>\d{1,2}|v\d{1,2})\s*[,\-:]?\s*(?P<title>[^\n\r]{3,160})",
        re.IGNORECASE,
    ),
    re.compile(
        r"(?P<title>[A-Z][A-Za-z0-9\s()&.,\-/]{3,140})\s*[,\-:]\s*(?P<doc>[A-Za-z]{1,10}[A-Za-z0-9\-_/]{1,30})\s*(?:rev(?:ision)?\s*)?(?P<rev>\d{1,2}|v\d{1,2})",
        re.IGNORECASE,
    ),
]


def normalize_revision(rev: str) -> str:
    rev = rev.strip().lower()
    rev = rev.replace("version", "").replace("revision", "").replace("rev", "").strip()
    rev = rev.replace("v", "").strip()
    if rev.isdigit():
        return rev.zfill(2)
    return rev.upper()


def normalize_text_blocks(text: str) -> str:
    # Normalize whitespace but keep line boundaries for heuristic splitting.
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
    return "\n".join([line for line in lines if line])


def extract_name(full_text: str) -> str:
    for pattern in NAME_PATTERNS:
        match = pattern.search(full_text)
        if match:
            return match.group(1).strip(" ,.-")
    return ""


def extract_employee_id(full_text: str) -> str:
    for pattern in EMPLOYEE_ID_PATTERNS:
        match = pattern.search(full_text)
        if match:
            return match.group(1).strip()
    return ""


def looks_like_header_cell(cell: str, candidates: Iterable[str]) -> bool:
    cell_low = (cell or "").strip().lower()
    return any(token in cell_low for token in candidates)


def parse_table_documents(table: List[List[Optional[str]]]) -> List[Tuple[str, str]]:
    if not table:
        return []

    header_idx = None
    doc_col = None
    title_col = None

    for i, row in enumerate(table[:4]):
        for j, cell in enumerate(row):
            if not cell:
                continue
            if looks_like_header_cell(cell, ["document number", "doc no", "doc #"]):
                doc_col = j
                header_idx = i
            if looks_like_header_cell(cell, ["description", "topic", "title"]):
                title_col = j
                header_idx = i if header_idx is None else header_idx

    if header_idx is None:
        return []

    results: List[Tuple[str, str]] = []
    for row in table[header_idx + 1 :]:
        if not row:
            continue
        doc_text = (row[doc_col] if doc_col is not None and doc_col < len(row) else "") or ""
        title_text = (row[title_col] if title_col is not None and title_col < len(row) else "") or ""

        doc_text = re.sub(r"\s+", " ", doc_text).strip()
        title_text = re.sub(r"\s+", " ", title_text).strip()
        if not doc_text and not title_text:
            continue

        # Handle line-break separated entries inside one table cell.
        doc_lines = [x.strip() for x in re.split(r"\n|;", doc_text) if x.strip()] or [doc_text]
        title_lines = [x.strip() for x in re.split(r"\n|;", title_text) if x.strip()] or [title_text]
        max_len = max(len(doc_lines), len(title_lines))

        for i in range(max_len):
            d = doc_lines[i] if i < len(doc_lines) else ""
            t = title_lines[i] if i < len(title_lines) else title_lines[-1]
            rev = ""
            m = re.search(r"(?:rev(?:ision)?\s*)?(v?\d{1,2})\b", d, re.IGNORECASE)
            if m:
                rev = normalize_revision(m.group(1))
            if d or t:
                title = f"{d}, {t}".strip(" ,") if d and t else (d or t)
                results.append((title, rev))

    return results


def parse_text_documents(full_text: str) -> List[Tuple[str, str]]:
    results: List[Tuple[str, str]] = []

    # Primary pattern extraction.
    for pattern in DOC_REV_PATTERNS:
        for match in pattern.finditer(full_text):
            doc = re.sub(r"\s+", " ", match.group("doc")).strip(" ,.-")
            rev = normalize_revision(match.group("rev"))
            title = re.sub(r"\s+", " ", match.group("title")).strip(" ,.-")
            if len(title) < 3:
                continue
            results.append((f"{doc}, {title}", rev))

    # Secondary fallback for broken lines: line that ends with revision, followed by title line.
    lines = [line.strip() for line in full_text.splitlines() if line.strip()]
    for idx in range(len(lines) - 1):
        m = re.search(r"^([A-Za-z]{1,10}[A-Za-z0-9\-_/]{1,30}).*?(?:rev(?:ision)?\s*)?(v?\d{1,2})\s*$", lines[idx], re.IGNORECASE)
        if not m:
            continue
        next_line = lines[idx + 1]
        if len(next_line) < 3:
            continue
        doc = m.group(1)
        rev = normalize_revision(m.group(2))
        results.append((f"{doc}, {next_line}", rev))

    # Deduplicate while preserving order.
    dedup: List[Tuple[str, str]] = []
    seen = set()
    for title, rev in results:
        key = (title.lower(), rev)
        if key in seen:
            continue
        seen.add(key)
        dedup.append((title, rev))

    return dedup


def _parse_doc_entry(entry: str) -> Optional[Tuple[str, str]]:
    """Parse a single document entry line such as
    'QS-03 Human Resources Management Procedure, version 02'
    and return (full_title, revision) or None if the line is not a document entry.
    """
    entry = entry.strip()
    if not entry:
        return None

    # Match "version NN" or "Rev NN" / "Revision NN" at the end (optionally preceded by comma)
    rev_match = re.search(r",?\s*(?:version|rev(?:ision)?)\s+(\d{1,2})\b.*$", entry, re.IGNORECASE)
    if rev_match:
        revision = rev_match.group(1).zfill(2)
        title = entry[: rev_match.start()].strip().rstrip(",").strip()
    else:
        # No revision info found – skip lines that don't look like document entries
        doc_prefix = re.match(r"^[A-Za-z]{1,10}[-/][A-Za-z0-9]{1,10}\b", entry)
        if not doc_prefix:
            return None
        revision = ""
        title = entry

    if not title:
        return None
    return (title, revision)


def parse_group_training_record(tables: List[List[List[Optional[str]]]]) -> List[Record]:
    """Parse the NUS Life Sciences Institute Group Training Record PDF format.

    The form has one large table with two logical sections:
    1. A document block at the top, identified by a header cell containing
       "Document Number" (and possibly "Revision").  Document entries may appear
       one per row or as multiple newline-separated lines inside a single cell.
    2. A trainee roster below, identified by a header row whose cells contain
       "Trainee Name" and "Trainee Employee ID".

    Returns the cross-product of trainees × documents as ``List[Record]``.
    Returns an empty list when the format is not recognised (so the caller can
    fall back to the existing general-purpose parser).
    """
    documents: List[Tuple[str, str]] = []  # (title, revision)
    trainees: List[Tuple[str, str]] = []  # (name, employee_id)

    for table in tables:
        if not table:
            continue

        doc_block_start: Optional[int] = None
        doc_col: Optional[int] = None
        trainee_header_row: Optional[int] = None
        trainee_name_col: Optional[int] = None
        trainee_id_col: Optional[int] = None

        for i, row in enumerate(table):
            for j, cell in enumerate(row):
                if not cell:
                    continue
                cell_low = cell.strip().lower()
                if "document number" in cell_low and doc_block_start is None:
                    doc_block_start = i
                    doc_col = j
                if "trainee name" in cell_low and trainee_header_row is None:
                    trainee_header_row = i
                    trainee_name_col = j
                if trainee_header_row == i and ("trainee employee id" in cell_low or (
                        "employee id" in cell_low and "trainee" in cell_low)):
                    trainee_id_col = j

        # Extract documents from the document block
        if doc_block_start is not None and doc_col is not None:
            doc_end = trainee_header_row if trainee_header_row is not None else len(table)
            for row in table[doc_block_start + 1 : doc_end]:
                if not row or doc_col >= len(row):
                    continue
                cell = (row[doc_col] or "").strip()
                if not cell:
                    continue
                for line in cell.splitlines():
                    parsed = _parse_doc_entry(line)
                    if parsed:
                        documents.append(parsed)

        # Extract trainees from the trainee roster
        if trainee_header_row is not None:
            for row in table[trainee_header_row + 1 :]:
                if not row:
                    continue
                name = ""
                emp_id = ""
                if trainee_name_col is not None and trainee_name_col < len(row):
                    name = (row[trainee_name_col] or "").strip()
                if trainee_id_col is not None and trainee_id_col < len(row):
                    emp_id = (row[trainee_id_col] or "").strip()
                # Skip blank or placeholder rows
                if not name or re.match(r"^[\s_\-]*$", name):
                    continue
                trainees.append((name, emp_id))

    if not documents or not trainees:
        return []

    # Cross-join: every trainee × every document → one Record
    return [
        Record(name=name, employee_id=emp_id, title=title, revision=rev)
        for name, emp_id in trainees
        for title, rev in documents
    ]


def extract_records_from_pdf(file_bytes: bytes, source_name: str) -> List[Record]:
    all_text_parts: List[str] = []
    table_documents: List[Tuple[str, str]] = []
    all_tables: List[List[List[Optional[str]]]] = []

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            all_text_parts.append(page_text)
            tables = page.extract_tables() or []
            all_tables.extend(tables)
            for table in tables:
                table_documents.extend(parse_table_documents(table))

    # Try the NUS Group Training Record format first
    group_records = parse_group_training_record(all_tables)
    if group_records:
        return group_records

    full_text = normalize_text_blocks("\n".join(all_text_parts))
    name = extract_name(full_text)
    employee_id = extract_employee_id(full_text)

    docs = table_documents if table_documents else parse_text_documents(full_text)

    # If parsing found no documents, return one row with metadata to help user review manually.
    if not docs:
        return [
            Record(
                name=name or "Unknown",
                employee_id=employee_id or "Unknown",
                title=f"Unable to confidently parse document list from {source_name}",
                revision="",
            )
        ]

    return [
        Record(
            name=name or "Unknown",
            employee_id=employee_id or "Unknown",
            title=title,
            revision=rev,
        )
        for title, rev in docs
    ]


def build_excel(records: List[Record]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Trainer Records"

    # Header row
    ws.merge_cells("A1:D1")
    ws["A1"] = "trainer records"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Column headers
    headers = ["Name", "Employee ID", "Title", "Revision"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for record in records:
        ws.append([record.name, record.employee_id, record.title, record.revision])

    # Basic sizing for readability.
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 12

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()


def records_to_dataframe(records: List[Record]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Name": r.name,
                "Employee ID": r.employee_id,
                "Title": r.title,
                "Revision": r.revision,
            }
            for r in records
        ]
    )


def main() -> None:
    st.set_page_config(page_title="Training Record PDF to Excel", page_icon="📄", layout="wide")
    st.title("Training Record PDF to Excel")
    st.write(
        "Upload one or more training-record PDFs. "
        "The app extracts trainee name, trainee employee ID, document title, and revision, "
        "then generates an Excel file with the requested format."
    )

    uploaded_files = st.file_uploader(
        "Upload PDF files",
        type=["pdf"],
        accept_multiple_files=True,
        help="You can upload multiple training records at once.",
    )

    if not uploaded_files:
        st.info("Upload at least one PDF to begin.")
        return

    all_records: List[Record] = []
    with st.spinner("Parsing PDF files..."):
        for file in uploaded_files:
            try:
                records = extract_records_from_pdf(file.read(), file.name)
                all_records.extend(records)
            except Exception as exc:
                st.warning(f"Could not parse {file.name}: {exc}")

    if not all_records:
        st.error("No records could be extracted from the uploaded files.")
        return

    df = records_to_dataframe(all_records)
    st.subheader("Extracted Records Preview")
    st.dataframe(df, use_container_width=True)

    excel_bytes = build_excel(all_records)
    st.download_button(
        label="Download Excel",
        data=excel_bytes,
        file_name="trainer_records.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.caption(
        "Note: For scanned PDFs or unusual layouts, extraction may need review. "
        "The app includes fallback parsing but cannot guarantee 100% accuracy for all templates."
    )


if __name__ == "__main__":
    main()
